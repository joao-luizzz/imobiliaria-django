from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.http import require_http_methods, require_POST
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.core.paginator import Paginator
from django.http import HttpResponse
from .models import Simulation, Cliente, MetaCorretor, AuditLog, UserProfile
from .calculos import calcular_sac, calcular_price
from collections import defaultdict
from functools import wraps
from io import BytesIO
import json
import datetime
import uuid
import math
import urllib.request


def staff_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.conf import settings
            return redirect(settings.LOGIN_URL)
        if not request.user.is_staff:
            messages.error(request, 'Acesso restrito a administradores.')
            return redirect('simulador:simular')
        return view_func(request, *args, **kwargs)
    return wrapper


@login_required
@require_http_methods(["GET", "POST"])
def simular(request):
    resultado = None

    if request.method == "POST":
        try:
            valor_imovel = float(request.POST.get("valor_imovel", 0))
            valor_entrada = float(request.POST.get("entrada", 0))
            taxa_juros = float(request.POST.get("taxa_juros", 0))
            prazo_meses = int(request.POST.get("meses", 0))
            sistema = request.POST.get("sistema", "SAC").upper()
            cliente = request.POST.get("cliente", "").strip()
            observacoes = request.POST.get("observacoes", "").strip()
            mip_mensal = float(request.POST.get("mip_mensal", 0) or 0)
            dfi_mensal = float(request.POST.get("dfi_mensal", 0) or 0)
            tags = request.POST.get("tags", "").strip()

            if valor_entrada >= valor_imovel:
                raise ValueError("A entrada não pode ser maior ou igual ao valor do imóvel.")
            if taxa_juros <= 0 or prazo_meses <= 0:
                raise ValueError("Taxa de juros e prazo devem ser maiores que zero.")
            if not cliente:
                raise ValueError("O nome do cliente é obrigatório.")

            valor_financiado = valor_imovel - valor_entrada

            if sistema == "SAC":
                parcelas = calcular_sac(valor_financiado, taxa_juros, prazo_meses, mip_mensal, dfi_mensal, valor_imovel)
            else:
                parcelas = calcular_price(valor_financiado, taxa_juros, prazo_meses, mip_mensal, dfi_mensal, valor_imovel)

            sim = Simulation.objects.create(
                usuario=request.user,
                cliente=cliente,
                valor_imovel=valor_imovel,
                entrada=valor_entrada,
                taxa_juros=taxa_juros,
                prazo_meses=prazo_meses,
                sistema=sistema,
                mip_mensal=mip_mensal,
                dfi_mensal=dfi_mensal,
                observacoes=observacoes,
                tags=tags,
            )
            registrar_log(request, "Criou simulação", "Simulation", sim.pk,
                          f"{cliente} — R$ {valor_imovel:.0f} ({sistema})")

            total_pago = sum(float(p["valor"]) for p in parcelas) if parcelas else 0
            total_juros = total_pago - valor_financiado
            total_seguros = sum(float(p["seguro"]) for p in parcelas) if parcelas else 0
            primeira_parcela = parcelas[0]["valor"] if parcelas else "0.00"
            ultima_parcela = parcelas[-1]["valor"] if parcelas else "0.00"

            # Dados para o gráfico de evolução do saldo devedor
            passo = max(1, prazo_meses // 60)
            indices = list(range(0, prazo_meses, passo))
            if parcelas and (prazo_meses - 1) not in indices:
                indices.append(prazo_meses - 1)
            chart_labels = [parcelas[i]['numero'] for i in indices]
            chart_saldo = [float(parcelas[i]['saldo_devedor']) for i in indices]

            resultado = {
                "sistema": sistema,
                "valor_imovel": f"{valor_imovel:.2f}",
                "entrada": f"{valor_entrada:.2f}",
                "valor_financiado": f"{valor_financiado:.2f}",
                "taxa_juros": f"{taxa_juros:.2f}",
                "meses": prazo_meses,
                "primeira_parcela": primeira_parcela,
                "ultima_parcela": ultima_parcela,
                "total_pago": f"{total_pago:.2f}",
                "total_juros": f"{total_juros:.2f}",
                "total_seguros": f"{total_seguros:.2f}",
                "custo_total": f"{valor_entrada + total_pago:.2f}",
                "usar_seguro": mip_mensal > 0 or dfi_mensal > 0,
                "mip_mensal": f"{mip_mensal:.4f}",
                "dfi_mensal": f"{dfi_mensal:.4f}",
                "parcelas": parcelas,
                "cliente": cliente,
                "chart_labels": json.dumps(chart_labels),
                "chart_saldo": json.dumps(chart_saldo),
            }

        except (ValueError, TypeError) as e:
            messages.error(request, f"Erro nos dados informados: {e}")
        except Exception as e:
            messages.error(request, f"Ocorreu um erro inesperado: {e}")

    return render(request, "simulador/index.html", {"resultado": resultado})


@login_required
def dashboard(request):
    # Admin vê tudo, outros usuários veem apenas suas próprias simulações
    if request.user.is_staff:
        qs = Simulation.objects.all()
    else:
        qs = Simulation.objects.filter(usuario=request.user)

    total = qs.count()
    volume = qs.aggregate(total=Sum('valor_imovel'))['total'] or 0
    ticket_medio = (volume / total) if total > 0 else 0

    # KPIs extras
    aprovados = qs.filter(status='aprovado').count()
    aprovados_pct = round(aprovados / total * 100, 1) if total > 0 else 0
    este_mes = qs.filter(
        criado_em__year=datetime.date.today().year,
        criado_em__month=datetime.date.today().month,
    ).count()
    sac_count = qs.filter(sistema='SAC').count()
    price_count = qs.filter(sistema='PRICE').count()

    # Gráfico de timeline: simulações por dia
    por_dia = defaultdict(int)
    for sim in qs.values('criado_em'):
        dia = sim['criado_em'].date().strftime('%d/%m/%Y')
        por_dia[dia] += 1

    dias_ordenados = sorted(por_dia.keys(), key=lambda d: d.split('/')[::-1])
    timeline_labels = dias_ordenados
    timeline_data = [por_dia[d] for d in dias_ordenados]

    # Gráfico de pizza: distribuição por status
    status_qs = qs.values('status').annotate(total=Count('status'))
    status_map = {s: label for s, label in Simulation.STATUS_CHOICES}
    status_labels = [status_map.get(s['status'], s['status']) for s in status_qs]
    status_data = [s['total'] for s in status_qs]

    # Atividade recente
    recentes = qs.select_related('usuario').order_by('-criado_em')[:5]

    context = {
        'total': total,
        'volume': volume,
        'ticket_medio': ticket_medio,
        'aprovados': aprovados,
        'aprovados_pct': aprovados_pct,
        'este_mes': este_mes,
        'sac_count': sac_count,
        'price_count': price_count,
        'timeline_labels': json.dumps(timeline_labels),
        'timeline_data': json.dumps(timeline_data),
        'status_labels': json.dumps(status_labels),
        'status_data': json.dumps(status_data),
        'sistema_labels': json.dumps(['SAC', 'PRICE']),
        'sistema_data': json.dumps([sac_count, price_count]),
        'recentes': recentes,
    }
    return render(request, 'simulador/dashboard.html', context)


@login_required
def historico(request):
    if request.user.is_staff:
        qs = Simulation.objects.select_related('usuario').all()
    else:
        qs = Simulation.objects.filter(usuario=request.user)

    # Filtros
    busca = request.GET.get('busca', '').strip()
    filtro_status = request.GET.get('status', '')
    filtro_sistema = request.GET.get('sistema', '')
    filtro_tag = request.GET.get('filtro_tag', '').strip()
    data_inicio = request.GET.get('data_inicio', '').strip()
    data_fim = request.GET.get('data_fim', '').strip()

    if busca:
        qs = qs.filter(cliente__icontains=busca)
    if filtro_status:
        qs = qs.filter(status=filtro_status)
    if filtro_sistema:
        qs = qs.filter(sistema=filtro_sistema)
    if filtro_tag:
        qs = qs.filter(tags__icontains=filtro_tag)
    if data_inicio:
        try:
            qs = qs.filter(criado_em__date__gte=datetime.date.fromisoformat(data_inicio))
        except ValueError:
            pass
    if data_fim:
        try:
            qs = qs.filter(criado_em__date__lte=datetime.date.fromisoformat(data_fim))
        except ValueError:
            pass

    # Paginação
    paginator = Paginator(qs, 10)
    page = request.GET.get('page', 1)
    simulacoes = paginator.get_page(page)

    context = {
        'simulacoes': simulacoes,
        'busca': busca,
        'filtro_status': filtro_status,
        'filtro_sistema': filtro_sistema,
        'filtro_tag': filtro_tag,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'status_choices': Simulation.STATUS_CHOICES,
        'sistema_choices': Simulation.SISTEMA_CHOICES,
        'total_filtrado': qs.count(),
    }
    return render(request, 'simulador/historico.html', context)


@login_required
@require_POST
def excluir_simulacao(request, pk):
    if request.user.is_staff:
        sim = get_object_or_404(Simulation, pk=pk)
    else:
        sim = get_object_or_404(Simulation, pk=pk, usuario=request.user)

    cliente = sim.cliente
    sim.delete()
    registrar_log(request, "Excluiu simulação", "Simulation", pk, cliente)
    messages.success(request, f'Simulação de "{cliente}" excluída com sucesso.')
    return redirect('simulador:historico')


@login_required
@require_http_methods(["GET", "POST"])
def editar_simulacao(request, pk):
    if request.user.is_staff:
        sim = get_object_or_404(Simulation, pk=pk)
    else:
        sim = get_object_or_404(Simulation, pk=pk, usuario=request.user)

    if request.method == 'POST':
        try:
            valor_imovel = float(request.POST.get('valor_imovel', 0))
            valor_entrada = float(request.POST.get('entrada', 0))
            taxa_juros = float(request.POST.get('taxa_juros', 0))
            prazo_meses = int(request.POST.get('meses', 0))
            sistema = request.POST.get('sistema', 'SAC').upper()
            cliente = request.POST.get('cliente', '').strip()
            observacoes = request.POST.get('observacoes', '').strip()

            if valor_entrada >= valor_imovel:
                raise ValueError("A entrada não pode ser maior ou igual ao valor do imóvel.")
            if taxa_juros <= 0 or prazo_meses <= 0:
                raise ValueError("Taxa de juros e prazo devem ser maiores que zero.")
            if not cliente:
                raise ValueError("O nome do cliente é obrigatório.")

            sim.cliente = cliente
            sim.valor_imovel = valor_imovel
            sim.entrada = valor_entrada
            sim.taxa_juros = taxa_juros
            sim.prazo_meses = prazo_meses
            sim.sistema = sistema
            sim.mip_mensal = float(request.POST.get('mip_mensal', 0) or 0)
            sim.dfi_mensal = float(request.POST.get('dfi_mensal', 0) or 0)
            sim.observacoes = observacoes
            sim.tags = request.POST.get('tags', '').strip()
            sim.save()
            registrar_log(request, "Editou simulação", "Simulation", sim.pk, cliente)
            messages.success(request, 'Simulação atualizada com sucesso.')
            return redirect('simulador:detalhe_simulacao', pk=sim.pk)
        except (ValueError, TypeError) as e:
            messages.error(request, f'Erro nos dados informados: {e}')

    return render(request, 'simulador/editar_simulacao.html', {'sim': sim})


@login_required
@require_POST
def toggle_favorito(request, pk):
    if request.user.is_staff:
        sim = get_object_or_404(Simulation, pk=pk)
    else:
        sim = get_object_or_404(Simulation, pk=pk, usuario=request.user)
    sim.favorito = not sim.favorito
    sim.save(update_fields=['favorito'])
    from django.http import JsonResponse
    return JsonResponse({'favorito': sim.favorito})


@login_required
@require_POST
def alterar_status(request, pk):
    if request.user.is_staff:
        sim = get_object_or_404(Simulation, pk=pk)
    else:
        sim = get_object_or_404(Simulation, pk=pk, usuario=request.user)
    novo_status = request.POST.get('status', '')
    status_validos = [s for s, _ in Simulation.STATUS_CHOICES]
    if novo_status in status_validos:
        sim.status = novo_status
        sim.save(update_fields=['status'])
        registrar_log(request, "Alterou status", "Simulation", pk,
                      f"{sim.cliente}: {sim.get_status_display()}")
        messages.success(request, f'Status de "{sim.cliente}" atualizado para "{sim.get_status_display()}".')
    else:
        messages.error(request, 'Status inválido.')

    return redirect('simulador:historico')


@login_required
def oraculo(request):
    resultado = None

    if request.method == 'POST':
        try:
            renda = float(request.POST.get('renda', 0))
            entrada = float(request.POST.get('entrada', 0))
            prazo_anos = int(request.POST.get('prazo_anos', 30))
            taxa_anual = float(request.POST.get('taxa_anual', 9.99))
            comprometimento = float(request.POST.get('comprometimento', 30))

            if renda <= 0:
                raise ValueError("A renda mensal deve ser maior que zero.")

            # Cálculo reverso: PMT -> PV (Price)
            margem_parcela = renda * (comprometimento / 100)
            taxa_mensal = (taxa_anual / 100) / 12
            meses = prazo_anos * 12
            valor_financiavel = margem_parcela * ((1 - (1 + taxa_mensal) ** (-meses)) / taxa_mensal)
            poder_compra = valor_financiavel + entrada

            def _brl(n):
                return f'{n:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')

            resultado = {
                'poder_compra': f'{poder_compra:,.2f}',
                'valor_financiavel': f'{valor_financiavel:,.2f}',
                'margem_parcela': f'{margem_parcela:,.2f}',
                'renda': f'{renda:,.2f}',
                'entrada': f'{entrada:,.2f}',
                'prazo_anos': prazo_anos,
                'taxa_anual': taxa_anual,
                'comprometimento': comprometimento,
                'interpretacao': (
                    f'Com renda de R$ {_brl(renda)}, o banco libera uma parcela de até '
                    f'R$ {_brl(margem_parcela)} ({comprometimento:.0f}% da renda). '
                    f'Isso financia até R$ {_brl(valor_financiavel)} em {prazo_anos} anos, '
                    f'dando um poder de compra de R$ {_brl(poder_compra)}.'
                ),
            }
        except (ValueError, TypeError) as e:
            messages.error(request, f'Erro nos dados informados: {e}')
        except Exception as e:
            messages.error(request, f'Ocorreu um erro inesperado: {e}')

    return render(request, 'simulador/oraculo.html', {'resultado': resultado})


@login_required
@require_http_methods(["GET", "POST"])
def comparativo(request):
    resultado = None

    if request.method == 'POST':
        try:
            valor_imovel = float(request.POST.get('valor_imovel', 0))
            valor_entrada = float(request.POST.get('entrada', 0))
            taxa_juros = float(request.POST.get('taxa_juros', 0))
            prazo_meses = int(request.POST.get('meses', 0))

            if valor_entrada >= valor_imovel:
                raise ValueError("A entrada não pode ser maior ou igual ao valor do imóvel.")
            if taxa_juros <= 0 or prazo_meses <= 0:
                raise ValueError("Taxa de juros e prazo devem ser maiores que zero.")

            valor_financiado = valor_imovel - valor_entrada
            parcelas_sac = calcular_sac(valor_financiado, taxa_juros, prazo_meses)
            parcelas_price = calcular_price(valor_financiado, taxa_juros, prazo_meses)

            def resumo(parcelas):
                total = sum(float(p['valor']) for p in parcelas)
                return {
                    'primeira': parcelas[0]['valor'],
                    'ultima': parcelas[-1]['valor'],
                    'total_pago': f'{total:,.2f}',
                    'total_juros': f'{total - valor_financiado:,.2f}',
                }

            # Amostragem do saldo devedor para o gráfico (máx 60 pontos)
            passo = max(1, prazo_meses // 60)
            indices = list(range(0, prazo_meses, passo))
            if (prazo_meses - 1) not in indices:
                indices.append(prazo_meses - 1)

            labels = [parcelas_sac[i]['numero'] for i in indices]
            saldo_sac = [float(parcelas_sac[i]['saldo_devedor']) for i in indices]
            saldo_price = [float(parcelas_price[i]['saldo_devedor']) for i in indices]

            resultado = {
                'sac': resumo(parcelas_sac),
                'price': resumo(parcelas_price),
                'valor_financiado': f'{valor_financiado:,.2f}',
                'labels': json.dumps(labels),
                'saldo_sac': json.dumps(saldo_sac),
                'saldo_price': json.dumps(saldo_price),
            }
        except (ValueError, TypeError) as e:
            messages.error(request, f'Erro nos dados informados: {e}')

    return render(request, 'simulador/comparativo.html', {'resultado': resultado})


@login_required
@require_http_methods(["GET", "POST"])
def amortizacao_extra(request):
    resultado = None
    if request.method == 'POST':
        try:
            valor_financiado = float(request.POST.get('valor_financiado', 0))
            taxa_juros = float(request.POST.get('taxa_juros', 0))
            prazo_meses = int(request.POST.get('meses', 0))
            sistema = request.POST.get('sistema', 'SAC').upper()
            aporte_extra = float(request.POST.get('aporte_extra', 0))

            if valor_financiado <= 0 or taxa_juros <= 0 or prazo_meses <= 0:
                raise ValueError("Preencha todos os campos com valores válidos.")
            if aporte_extra < 0:
                raise ValueError("O aporte extra não pode ser negativo.")

            taxa_m = taxa_juros / 100

            # ── Cenário base ──────────────────────────────────────────────
            parcelas_base = (calcular_sac if sistema == 'SAC' else calcular_price)(
                valor_financiado, taxa_juros, prazo_meses)
            total_juros_base = sum(float(p['juros']) for p in parcelas_base)

            # ── Cenário com aporte extra (mês a mês) ──────────────────────
            saldo = valor_financiado
            mes = 0
            total_juros_extra = 0.0

            if sistema == 'SAC':
                amort_base = valor_financiado / prazo_meses
                while saldo > 0.01:
                    juros = saldo * taxa_m
                    total_juros_extra += juros
                    amort = min(amort_base + aporte_extra, saldo)
                    saldo -= amort
                    mes += 1
                    if mes > prazo_meses * 2:
                        break
            else:
                if taxa_m > 0:
                    num = taxa_m * (1 + taxa_m) ** prazo_meses
                    den = (1 + taxa_m) ** prazo_meses - 1
                    pmt = valor_financiado * num / den
                else:
                    pmt = valor_financiado / prazo_meses
                while saldo > 0.01:
                    juros = saldo * taxa_m
                    total_juros_extra += juros
                    amort_price = pmt - juros
                    amort = min(amort_price + aporte_extra, saldo)
                    saldo -= amort
                    mes += 1
                    if mes > prazo_meses * 2:
                        break

            economia_juros = total_juros_base - total_juros_extra
            meses_economizados = prazo_meses - mes

            # Amostragem para gráfico (máx 60 pts) dos dois cenários
            passo = max(1, prazo_meses // 60)
            chart_labels = [p['numero'] for p in parcelas_base[::passo]]
            chart_base = [float(p['saldo_devedor']) for p in parcelas_base[::passo]]

            # recalcular saldos do cenário extra para o gráfico
            saldo2, saldos_extra = valor_financiado, []
            amort_base2 = valor_financiado / prazo_meses if sistema == 'SAC' else None
            if sistema == 'PRICE' and taxa_m > 0:
                num = taxa_m * (1 + taxa_m) ** prazo_meses
                pmt2 = valor_financiado * num / ((1 + taxa_m) ** prazo_meses - 1)
            elif sistema == 'PRICE':
                pmt2 = valor_financiado / prazo_meses

            for i in range(prazo_meses):
                if saldo2 <= 0:
                    saldos_extra.append(0)
                    continue
                j2 = saldo2 * taxa_m
                if sistema == 'SAC':
                    saldo2 -= min(amort_base2 + aporte_extra, saldo2)
                else:
                    saldo2 -= min((pmt2 - j2) + aporte_extra, saldo2)
                saldos_extra.append(round(saldo2, 2))

            chart_extra = saldos_extra[::passo]
            if len(chart_extra) < len(chart_labels):
                chart_extra += [0] * (len(chart_labels) - len(chart_extra))

            resultado = {
                'sistema': sistema,
                'valor_financiado': f'{valor_financiado:.2f}',
                'taxa_juros': f'{taxa_juros:.4f}',
                'meses': prazo_meses,
                'aporte_extra': f'{aporte_extra:.2f}',
                'meses_base': prazo_meses,
                'meses_extra': mes,
                'meses_economizados': meses_economizados,
                'total_juros_base': f'{total_juros_base:.2f}',
                'total_juros_extra': f'{total_juros_extra:.2f}',
                'economia_juros': f'{economia_juros:.2f}',
                'economia_pct': f'{(economia_juros / total_juros_base * 100):.1f}' if total_juros_base > 0 else '0',
                'chart_labels': json.dumps(chart_labels),
                'chart_base': json.dumps(chart_base),
                'chart_extra': json.dumps(chart_extra),
            }
        except (ValueError, TypeError) as e:
            messages.error(request, f'Erro: {e}')

    return render(request, 'simulador/amortizacao_extra.html', {'resultado': resultado})


@login_required
@require_http_methods(["GET", "POST"])
def portabilidade(request):
    resultado = None
    if request.method == 'POST':
        try:
            saldo_devedor = float(request.POST.get('saldo_devedor', 0))
            taxa_atual = float(request.POST.get('taxa_atual', 0))
            taxa_nova = float(request.POST.get('taxa_nova', 0))
            prazo_restante = int(request.POST.get('prazo_restante', 0))

            if saldo_devedor <= 0 or taxa_atual <= 0 or taxa_nova <= 0 or prazo_restante <= 0:
                raise ValueError("Preencha todos os campos com valores válidos.")

            parcelas_atual = calcular_price(saldo_devedor, taxa_atual, prazo_restante)
            parcelas_nova = calcular_price(saldo_devedor, taxa_nova, prazo_restante)

            total_atual = sum(float(p['valor']) for p in parcelas_atual)
            total_nova = sum(float(p['valor']) for p in parcelas_nova)
            juros_atual = sum(float(p['juros']) for p in parcelas_atual)
            juros_nova = sum(float(p['juros']) for p in parcelas_nova)

            parcela_atual = float(parcelas_atual[0]['valor'])
            parcela_nova = float(parcelas_nova[0]['valor'])
            economia_mensal = parcela_atual - parcela_nova
            economia_total = total_atual - total_nova

            passo = max(1, prazo_restante // 60)
            chart_labels = [p['numero'] for p in parcelas_atual[::passo]]
            chart_atual = [float(p['saldo_devedor']) for p in parcelas_atual[::passo]]
            chart_nova = [float(p['saldo_devedor']) for p in parcelas_nova[::passo]]

            resultado = {
                'saldo_devedor': f'{saldo_devedor:.2f}',
                'taxa_atual': f'{taxa_atual:.4f}',
                'taxa_nova': f'{taxa_nova:.4f}',
                'prazo_restante': prazo_restante,
                'parcela_atual': f'{parcela_atual:.2f}',
                'parcela_nova': f'{parcela_nova:.2f}',
                'economia_mensal': f'{economia_mensal:.2f}',
                'economia_total': f'{economia_total:.2f}',
                'total_atual': f'{total_atual:.2f}',
                'total_nova': f'{total_nova:.2f}',
                'juros_atual': f'{juros_atual:.2f}',
                'juros_nova': f'{juros_nova:.2f}',
                'economia_juros': f'{juros_atual - juros_nova:.2f}',
                'economia_pct': f'{(economia_total / total_atual * 100):.1f}' if total_atual > 0 else '0',
                'vale_porta': economia_mensal > 0,
                'chart_labels': json.dumps(chart_labels),
                'chart_atual': json.dumps(chart_atual),
                'chart_nova': json.dumps(chart_nova),
            }
        except (ValueError, TypeError) as e:
            messages.error(request, f'Erro: {e}')

    return render(request, 'simulador/portabilidade.html', {'resultado': resultado})


@login_required
@require_http_methods(["GET", "POST"])
def perfil(request):
    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')

        errors = []
        if password1:
            if password1 != password2:
                errors.append('As senhas não conferem.')
            elif len(password1) < 6:
                errors.append('A senha deve ter no mínimo 6 caracteres.')

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            request.user.first_name = first_name
            request.user.last_name = last_name
            request.user.email = email
            if password1:
                request.user.set_password(password1)
                from django.contrib.auth import update_session_auth_hash
                update_session_auth_hash(request, request.user)
            request.user.save()
            messages.success(request, 'Perfil atualizado com sucesso.')
            return redirect('simulador:perfil')

    return render(request, 'simulador/perfil.html', {
        'form_data': {
            'first_name': request.user.first_name,
            'last_name': request.user.last_name,
            'email': request.user.email,
        }
    })


@login_required
@require_POST
def gerar_link(request, pk):
    if request.user.is_staff:
        sim = get_object_or_404(Simulation, pk=pk)
    else:
        sim = get_object_or_404(Simulation, pk=pk, usuario=request.user)

    if not sim.share_token:
        sim.share_token = uuid.uuid4()
        sim.save(update_fields=['share_token'])

    link = request.build_absolute_uri(f'/s/{sim.share_token}/')
    from django.http import JsonResponse
    return JsonResponse({'link': link})


def simulacao_publica(request, token):
    sim = get_object_or_404(Simulation, share_token=token)
    fn = calcular_sac if sim.sistema == 'SAC' else calcular_price
    parcelas = fn(float(sim.valor_financiado), float(sim.taxa_juros), sim.prazo_meses,
                  float(sim.mip_mensal), float(sim.dfi_mensal), float(sim.valor_imovel))
    total_pago = sum(float(p['valor']) for p in parcelas)
    total_juros = total_pago - float(sim.valor_financiado)
    return render(request, 'simulador/simulacao_publica.html', {
        'sim': sim,
        'parcelas': parcelas,
        'total_pago': total_pago,
        'total_juros': total_juros,
        'primeira_parcela': parcelas[0]['valor'] if parcelas else '0.00',
        'ultima_parcela': parcelas[-1]['valor'] if parcelas else '0.00',
    })


@login_required
def exportar_pdf(request, pk):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    if request.user.is_staff:
        sim = get_object_or_404(Simulation, pk=pk)
    else:
        sim = get_object_or_404(Simulation, pk=pk, usuario=request.user)

    parcelas = (calcular_sac if sim.sistema == 'SAC' else calcular_price)(
        float(sim.valor_financiado), float(sim.taxa_juros), sim.prazo_meses,
        float(sim.mip_mensal), float(sim.dfi_mensal), float(sim.valor_imovel)
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    azul = colors.HexColor('#1a3c5e')
    azul_claro = colors.HexColor('#2e86c1')
    cinza = colors.HexColor('#f0f4f8')

    titulo_style = ParagraphStyle('titulo', parent=styles['Title'],
                                   textColor=azul, fontSize=18, spaceAfter=4)
    sub_style = ParagraphStyle('sub', parent=styles['Normal'],
                                textColor=azul_claro, fontSize=10, spaceAfter=12)
    label_style = ParagraphStyle('label', parent=styles['Normal'],
                                  textColor=colors.grey, fontSize=8)

    total_pago = sum(float(p['valor']) for p in parcelas)
    total_juros = total_pago - float(sim.valor_financiado)

    story = [
        Paragraph('Proposta de Financiamento', titulo_style),
        Paragraph(f'Emitido em {datetime.date.today().strftime("%d/%m/%Y")} — Sistema {sim.sistema}', sub_style),
        Spacer(1, 0.3*cm),
    ]

    # Dados resumo
    resumo = [
        ['Cliente', sim.cliente, 'Usuário', sim.usuario.get_full_name() or sim.usuario.username],
        ['Valor do Imóvel', f'R$ {float(sim.valor_imovel):,.2f}', 'Entrada', f'R$ {float(sim.entrada):,.2f}'],
        ['Valor Financiado', f'R$ {float(sim.valor_financiado):,.2f}', 'Taxa Mensal', f'{float(sim.taxa_juros):.2f}%'],
        ['Prazo', f'{sim.prazo_meses} meses', 'Total Pago', f'R$ {total_pago:,.2f}'],
        ['Total em Juros', f'R$ {total_juros:,.2f}', 'Status', sim.get_status_display()],
    ]
    t_resumo = Table(resumo, colWidths=[3.5*cm, 5.5*cm, 3.5*cm, 5.5*cm])
    t_resumo.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), cinza),
        ('BACKGROUND', (2, 0), (2, -1), cinza),
        ('TEXTCOLOR', (0, 0), (0, -1), azul),
        ('TEXTCOLOR', (2, 0), (2, -1), azul),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story += [t_resumo, Spacer(1, 0.5*cm)]

    # Tabela de parcelas (primeiras 24 + últimas 3)
    story.append(Paragraph('Tabela de Parcelas', ParagraphStyle(
        'h2', parent=styles['Heading2'], textColor=azul, fontSize=12, spaceAfter=6)))

    cabecalho = [['#', 'Parcela', 'Amortização', 'Juros', 'Saldo Devedor']]
    exibir = parcelas[:24] + (parcelas[-3:] if len(parcelas) > 27 else [])
    linhas = cabecalho + [
        [p['numero'], f'R$ {p["valor"]}', f'R$ {p["amortizacao"]}',
         f'R$ {p["juros"]}', f'R$ {p["saldo_devedor"]}']
        for p in exibir
    ]
    if len(parcelas) > 27:
        linhas.insert(25, ['...', '...', '...', '...', '...'])

    t_parcelas = Table(linhas, colWidths=[1.2*cm, 3.8*cm, 3.8*cm, 3.8*cm, 4.4*cm])
    t_parcelas.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), azul),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f7ff')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#dee2e6')),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(t_parcelas)

    doc.build(story)
    buffer.seek(0)
    nome = f'proposta_{sim.cliente.replace(" ", "_")}.pdf'
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nome}"'
    return response


@login_required
def exportar_excel(request, pk):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if request.user.is_staff:
        sim = get_object_or_404(Simulation, pk=pk)
    else:
        sim = get_object_or_404(Simulation, pk=pk, usuario=request.user)

    parcelas_sac = calcular_sac(float(sim.valor_financiado), float(sim.taxa_juros), sim.prazo_meses,
                                float(sim.mip_mensal), float(sim.dfi_mensal), float(sim.valor_imovel))
    parcelas_price = calcular_price(float(sim.valor_financiado), float(sim.taxa_juros), sim.prazo_meses,
                                    float(sim.mip_mensal), float(sim.dfi_mensal), float(sim.valor_imovel))

    wb = openpyxl.Workbook()
    azul = 'FF1A3C5E'
    azul_claro = 'FF2E86C1'
    cinza = 'FFF0F4F8'

    header_font = Font(name='Calibri', bold=True, color='FFFFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor=azul)
    sub_fill = PatternFill('solid', fgColor=azul_claro)
    center = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin', color='FFCFD8DD'),
        right=Side(style='thin', color='FFCFD8DD'),
        top=Side(style='thin', color='FFCFD8DD'),
        bottom=Side(style='thin', color='FFCFD8DD'),
    )

    def estilizar_cabecalho(ws, colunas):
        ws.append(colunas)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin

    def estilizar_linhas(ws, inicio=2):
        alt_fill = PatternFill('solid', fgColor='FFF0F7FF')
        for i, row in enumerate(ws.iter_rows(min_row=inicio)):
            for cell in row:
                cell.alignment = center
                cell.border = thin
                if i % 2 == 1:
                    cell.fill = alt_fill

    def ajustar_colunas(ws):
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col) + 4
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len, 25)

    # Aba Resumo
    ws_resumo = wb.active
    ws_resumo.title = 'Resumo'
    estilizar_cabecalho(ws_resumo, ['Campo', 'Valor'])
    dados_resumo = [
        ('Cliente', sim.cliente),
        ('Valor do Imóvel', f'R$ {float(sim.valor_imovel):,.2f}'),
        ('Entrada', f'R$ {float(sim.entrada):,.2f}'),
        ('Valor Financiado', f'R$ {float(sim.valor_financiado):,.2f}'),
        ('Taxa de Juros Mensal', f'{float(sim.taxa_juros):.2f}%'),
        ('Prazo', f'{sim.prazo_meses} meses'),
        ('Sistema', sim.sistema),
        ('Status', sim.get_status_display()),
        ('Data da Simulação', sim.criado_em.strftime('%d/%m/%Y %H:%M')),
    ]
    for row in dados_resumo:
        ws_resumo.append(row)
    estilizar_linhas(ws_resumo)
    ajustar_colunas(ws_resumo)

    # Aba SAC
    ws_sac = wb.create_sheet('Tabela SAC')
    estilizar_cabecalho(ws_sac, ['#', 'Parcela (R$)', 'Amortização (R$)', 'Juros (R$)', 'Saldo Devedor (R$)'])
    for p in parcelas_sac:
        ws_sac.append([p['numero'], p['valor'], p['amortizacao'], p['juros'], p['saldo_devedor']])
    estilizar_linhas(ws_sac)
    ajustar_colunas(ws_sac)

    # Aba PRICE
    ws_price = wb.create_sheet('Tabela PRICE')
    estilizar_cabecalho(ws_price, ['#', 'Parcela (R$)', 'Amortização (R$)', 'Juros (R$)', 'Saldo Devedor (R$)'])
    for p in parcelas_price:
        ws_price.append([p['numero'], p['valor'], p['amortizacao'], p['juros'], p['saldo_devedor']])
    estilizar_linhas(ws_price)
    ajustar_colunas(ws_price)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nome = f'comparativo_{sim.cliente.replace(" ", "_")}.xlsx'
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nome}"'
    return response


@login_required
def exportar_historico(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if request.user.is_staff:
        qs = Simulation.objects.select_related('usuario').all()
    else:
        qs = Simulation.objects.filter(usuario=request.user)

    busca = request.GET.get('busca', '').strip()
    filtro_status = request.GET.get('status', '')
    filtro_sistema = request.GET.get('sistema', '')
    data_inicio = request.GET.get('data_inicio', '').strip()
    data_fim = request.GET.get('data_fim', '').strip()

    if busca:
        qs = qs.filter(cliente__icontains=busca)
    if filtro_status:
        qs = qs.filter(status=filtro_status)
    if filtro_sistema:
        qs = qs.filter(sistema=filtro_sistema)
    if data_inicio:
        try:
            qs = qs.filter(criado_em__date__gte=datetime.date.fromisoformat(data_inicio))
        except ValueError:
            pass
    if data_fim:
        try:
            qs = qs.filter(criado_em__date__lte=datetime.date.fromisoformat(data_fim))
        except ValueError:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Histórico'

    azul = 'FF1A3C5E'
    header_font = Font(name='Calibri', bold=True, color='FFFFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor=azul)
    center = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin', color='FFCFD8DD'),
        right=Side(style='thin', color='FFCFD8DD'),
        top=Side(style='thin', color='FFCFD8DD'),
        bottom=Side(style='thin', color='FFCFD8DD'),
    )

    colunas = ['#', 'Cliente']
    if request.user.is_staff:
        colunas.append('Usuário')
    colunas += ['Valor Imóvel (R$)', 'Entrada (R$)', 'Valor Financiado (R$)',
                'Taxa Mensal (%)', 'Prazo (meses)', 'Sistema', 'Status', 'Data']

    ws.append(colunas)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin

    alt_fill = PatternFill('solid', fgColor='FFF0F7FF')
    status_map = dict(Simulation.STATUS_CHOICES)

    for i, sim in enumerate(qs.order_by('-criado_em'), start=1):
        row = [sim.pk, sim.cliente]
        if request.user.is_staff:
            row.append(sim.usuario.username)
        row += [
            float(sim.valor_imovel),
            float(sim.entrada),
            float(sim.valor_financiado),
            float(sim.taxa_juros),
            sim.prazo_meses,
            sim.sistema,
            status_map.get(sim.status, sim.status),
            sim.criado_em.strftime('%d/%m/%Y %H:%M'),
        ]
        ws.append(row)
        for cell in ws[i + 1]:
            cell.alignment = center
            cell.border = thin
            if i % 2 == 0:
                cell.fill = alt_fill

    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col) + 4
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len, 30)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filtros = '_'.join(filter(None, [busca, filtro_status, filtro_sistema])) or 'completo'
    nome = f'historico_{filtros}_{datetime.date.today().strftime("%Y%m%d")}.xlsx'
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nome}"'
    return response


@login_required
def detalhe_simulacao(request, pk):
    if request.user.is_staff:
        sim = get_object_or_404(Simulation, pk=pk)
    else:
        sim = get_object_or_404(Simulation, pk=pk, usuario=request.user)

    fn = calcular_sac if sim.sistema == 'SAC' else calcular_price
    parcelas = fn(float(sim.valor_financiado), float(sim.taxa_juros), sim.prazo_meses,
                  float(sim.mip_mensal), float(sim.dfi_mensal), float(sim.valor_imovel))
    total_pago = sum(float(p['valor']) for p in parcelas)
    total_juros = total_pago - float(sim.valor_financiado)
    total_seguros = sum(float(p['seguro']) for p in parcelas)

    # Gráfico comparativo SAC vs PRICE
    vf = float(sim.valor_financiado)
    tj = float(sim.taxa_juros)
    pm = sim.prazo_meses
    p_sac   = calcular_sac(vf, tj, pm)
    p_price = calcular_price(vf, tj, pm)
    passo = max(1, pm // 60)
    chart_labels     = [p['numero'] for p in p_sac[::passo]]
    chart_saldo_sac  = [float(p['saldo_devedor']) for p in p_sac[::passo]]
    chart_saldo_price= [float(p['saldo_devedor']) for p in p_price[::passo]]
    total_sac   = sum(float(p['valor']) for p in p_sac)
    total_price = sum(float(p['valor']) for p in p_price)
    juros_sac   = total_sac   - vf
    juros_price = total_price - vf

    return render(request, 'simulador/detalhe.html', {
        'sim': sim,
        'parcelas': parcelas,
        'total_pago': total_pago,
        'total_juros': total_juros,
        'total_seguros': total_seguros,
        'primeira_parcela': parcelas[0]['valor'] if parcelas else '0.00',
        'ultima_parcela': parcelas[-1]['valor'] if parcelas else '0.00',
        'chart_labels': json.dumps(chart_labels),
        'chart_saldo_sac': json.dumps(chart_saldo_sac),
        'chart_saldo_price': json.dumps(chart_saldo_price),
        'total_sac': f'{total_sac:.2f}',
        'total_price': f'{total_price:.2f}',
        'juros_sac': f'{juros_sac:.2f}',
        'juros_price': f'{juros_price:.2f}',
        'primeira_sac': p_sac[0]['valor'] if p_sac else '0.00',
        'primeira_price': p_price[0]['valor'] if p_price else '0.00',
    })



@login_required
@require_http_methods(["GET", "POST"])
def fgts(request):
    resultado = None
    if request.method == 'POST':
        try:
            valor_financiado = float(request.POST.get('valor_financiado', 0))
            taxa_juros = float(request.POST.get('taxa_juros', 0))
            prazo_meses = int(request.POST.get('meses', 0))
            sistema = request.POST.get('sistema', 'SAC').upper()
            saldo_fgts = float(request.POST.get('saldo_fgts', 0))
            modalidade = request.POST.get('modalidade', 'parcela')

            if any(v <= 0 for v in [valor_financiado, taxa_juros, prazo_meses, saldo_fgts]):
                raise ValueError("Preencha todos os campos com valores positivos.")
            if saldo_fgts >= valor_financiado:
                raise ValueError("O FGTS não pode ser maior ou igual ao valor financiado.")

            fn = calcular_sac if sistema == 'SAC' else calcular_price
            parcelas_base = fn(valor_financiado, taxa_juros, prazo_meses)
            total_base  = sum(float(p['valor']) for p in parcelas_base)
            juros_base  = sum(float(p['juros']) for p in parcelas_base)
            parcela_1_base = float(parcelas_base[0]['valor'])

            novo_financiado = valor_financiado - saldo_fgts
            taxa_m = taxa_juros / 100

            if modalidade == 'parcela':
                # Mesma prazo, principal reduzido → parcela menor
                parcelas_fgts = fn(novo_financiado, taxa_juros, prazo_meses)
                prazo_fgts = prazo_meses
                meses_economizados = 0
            else:
                # Mesmo PMT, prazo reduzido
                if sistema == 'PRICE':
                    pmt = float(parcelas_base[0]['valor'])
                    if taxa_m > 0 and pmt > novo_financiado * taxa_m:
                        novo_n = math.ceil(
                            -math.log(1 - novo_financiado * taxa_m / pmt) / math.log(1 + taxa_m)
                        )
                    else:
                        novo_n = prazo_meses
                else:
                    amort_base = valor_financiado / prazo_meses
                    novo_n = math.ceil(novo_financiado / amort_base)

                novo_n = max(1, min(novo_n, prazo_meses))
                parcelas_fgts = fn(novo_financiado, taxa_juros, novo_n)
                prazo_fgts = novo_n
                meses_economizados = prazo_meses - novo_n

            total_fgts = sum(float(p['valor']) for p in parcelas_fgts)
            juros_fgts = sum(float(p['juros']) for p in parcelas_fgts)
            parcela_1_fgts = float(parcelas_fgts[0]['valor'])
            economia_parcela = parcela_1_base - parcela_1_fgts
            economia_total   = total_base - total_fgts
            economia_juros   = juros_base - juros_fgts

            passo = max(1, prazo_meses // 60)
            chart_labels = [p['numero'] for p in parcelas_base[::passo]]
            chart_base   = [float(p['saldo_devedor']) for p in parcelas_base[::passo]]
            chart_fgts   = [float(p['saldo_devedor']) for p in parcelas_fgts[::passo]]
            if len(chart_fgts) < len(chart_labels):
                chart_fgts += [0] * (len(chart_labels) - len(chart_fgts))

            resultado = {
                'modalidade': modalidade,
                'sistema': sistema,
                'valor_financiado': f'{valor_financiado:.2f}',
                'novo_financiado': f'{novo_financiado:.2f}',
                'saldo_fgts': f'{saldo_fgts:.2f}',
                'taxa_juros': f'{taxa_juros:.4f}',
                'meses': prazo_meses,
                'parcela_base': f'{parcela_1_base:.2f}',
                'parcela_fgts': f'{parcela_1_fgts:.2f}',
                'prazo_base': prazo_meses,
                'prazo_fgts': prazo_fgts,
                'meses_economizados': meses_economizados,
                'economia_parcela': f'{economia_parcela:.2f}',
                'total_base': f'{total_base:.2f}',
                'total_fgts': f'{total_fgts:.2f}',
                'juros_base': f'{juros_base:.2f}',
                'juros_fgts': f'{juros_fgts:.2f}',
                'economia_total': f'{economia_total:.2f}',
                'economia_juros': f'{economia_juros:.2f}',
                'economia_pct': f'{(economia_total / total_base * 100):.1f}' if total_base > 0 else '0',
                'chart_labels': json.dumps(chart_labels),
                'chart_base': json.dumps(chart_base),
                'chart_fgts': json.dumps(chart_fgts),
            }
        except (ValueError, TypeError) as e:
            messages.error(request, f'Erro: {e}')

    return render(request, 'simulador/fgts.html', {'resultado': resultado})


@login_required
@require_http_methods(["GET", "POST"])
def itbi(request):
    resultado = None
    if request.method == 'POST':
        try:
            valor_imovel     = float(request.POST.get('valor_imovel', 0))
            aliquota_itbi    = float(request.POST.get('aliquota_itbi', 2.0))
            cartorio_percent = float(request.POST.get('cartorio_percent', 1.0))
            avaliacao        = float(request.POST.get('avaliacao', 3000) or 3000)
            certidoes        = float(request.POST.get('certidoes', 500) or 500)

            if valor_imovel <= 0:
                raise ValueError("Informe o valor do imóvel.")

            itbi_val   = valor_imovel * (aliquota_itbi / 100)
            cartorio   = valor_imovel * (cartorio_percent / 100)
            total_taxas = itbi_val + cartorio + avaliacao + certidoes
            custo_total = valor_imovel + total_taxas
            pct         = (total_taxas / valor_imovel * 100) if valor_imovel > 0 else 0

            resultado = {
                'valor_imovel':     f'{valor_imovel:.2f}',
                'aliquota_itbi':    f'{aliquota_itbi:.2f}',
                'cartorio_percent': f'{cartorio_percent:.2f}',
                'itbi':             f'{itbi_val:.2f}',
                'cartorio':         f'{cartorio:.2f}',
                'avaliacao':        f'{avaliacao:.2f}',
                'certidoes':        f'{certidoes:.2f}',
                'total_taxas':      f'{total_taxas:.2f}',
                'custo_total':      f'{custo_total:.2f}',
                'pct_sobre_imovel': f'{pct:.1f}',
                # chart: breakdown
                'chart_labels': json.dumps(['ITBI', 'Cartório/Registro', 'Aval. Bancária', 'Certidões']),
                'chart_data':   json.dumps([round(itbi_val, 2), round(cartorio, 2),
                                            round(avaliacao, 2), round(certidoes, 2)]),
            }
        except (ValueError, TypeError) as e:
            messages.error(request, f'Erro: {e}')

    return render(request, 'simulador/itbi.html', {'resultado': resultado})


@login_required
@require_http_methods(["GET", "POST"])
def ipca_tr(request):
    resultado = None
    if request.method == 'POST':
        try:
            valor_financiado = float(request.POST.get('valor_financiado', 0))
            taxa_juros       = float(request.POST.get('taxa_juros', 0))
            prazo_meses      = int(request.POST.get('meses', 0))
            sistema          = request.POST.get('sistema', 'SAC').upper()
            taxa_correcao    = float(request.POST.get('taxa_correcao', 0))

            if valor_financiado <= 0 or taxa_juros <= 0 or prazo_meses <= 0:
                raise ValueError("Preencha todos os campos com valores válidos.")

            taxa_m = taxa_juros / 100
            corr_m = taxa_correcao / 100

            # Cenário base (sem correção)
            fn = calcular_sac if sistema == 'SAC' else calcular_price
            parcelas_base = fn(valor_financiado, taxa_juros, prazo_meses)
            total_base  = sum(float(p['valor']) for p in parcelas_base)
            juros_base  = sum(float(p['juros']) for p in parcelas_base)

            # Cenário com correção mensal (IPCA/TR sobre saldo)
            saldo = valor_financiado
            total_corr = 0.0
            juros_corr = 0.0
            saldos_corr = []

            if sistema == 'SAC':
                amort_base = valor_financiado / prazo_meses
                for _ in range(prazo_meses):
                    if saldo <= 0.01:
                        saldos_corr.append(0)
                        continue
                    saldo  *= (1 + corr_m)
                    juros   = saldo * taxa_m
                    amort   = min(amort_base * (1 + corr_m), saldo)
                    total_corr += amort + juros
                    juros_corr += juros
                    saldo -= amort
                    saldos_corr.append(round(max(saldo, 0), 2))
            else:
                pmt = float(parcelas_base[0]['valor'])
                for _ in range(prazo_meses):
                    if saldo <= 0.01:
                        saldos_corr.append(0)
                        continue
                    saldo *= (1 + corr_m)
                    juros  = saldo * taxa_m
                    amort  = max(pmt - juros, 0)
                    total_corr += pmt
                    juros_corr += juros
                    saldo  = max(saldo - amort, 0)
                    saldos_corr.append(round(saldo, 2))

            custo_extra = total_corr - total_base

            passo = max(1, prazo_meses // 60)
            chart_labels = [p['numero'] for p in parcelas_base[::passo]]
            chart_base   = [float(p['saldo_devedor']) for p in parcelas_base[::passo]]
            chart_corr   = saldos_corr[::passo]
            if len(chart_corr) < len(chart_labels):
                chart_corr += [0] * (len(chart_labels) - len(chart_corr))

            resultado = {
                'sistema': sistema,
                'valor_financiado': f'{valor_financiado:.2f}',
                'taxa_juros':       f'{taxa_juros:.4f}',
                'meses':            prazo_meses,
                'taxa_correcao':    f'{taxa_correcao:.4f}',
                'total_base':       f'{total_base:.2f}',
                'total_corr':       f'{total_corr:.2f}',
                'juros_base':       f'{juros_base:.2f}',
                'juros_corr':       f'{juros_corr:.2f}',
                'custo_extra':      f'{custo_extra:.2f}',
                'custo_extra_pct':  f'{(custo_extra / total_base * 100):.1f}' if total_base > 0 else '0',
                'chart_labels':     json.dumps(chart_labels),
                'chart_base':       json.dumps(chart_base),
                'chart_corr':       json.dumps(chart_corr),
            }
        except (ValueError, TypeError) as e:
            messages.error(request, f'Erro: {e}')

    return render(request, 'simulador/ipca_tr.html', {'resultado': resultado})


# ── Gestão de Usuários (staff only) ──────────────────────────────────────────

@staff_required
def usuarios_lista(request):
    usuarios = User.objects.all().order_by('username').annotate(
        total_sims=Count('simulacoes')
    )
    return render(request, 'simulador/usuarios_lista.html', {'usuarios': usuarios})


@staff_required
@require_http_methods(["GET", "POST"])
def usuario_criar(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')
        is_staff = request.POST.get('is_staff') == 'on'

        errors = []
        if not username:
            errors.append('O nome de usuário é obrigatório.')
        elif User.objects.filter(username=username).exists():
            errors.append(f'O usuário "{username}" já existe.')
        if not password1:
            errors.append('A senha é obrigatória.')
        elif password1 != password2:
            errors.append('As senhas não conferem.')
        elif len(password1) < 6:
            errors.append('A senha deve ter no mínimo 6 caracteres.')

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            User.objects.create_user(
                username=username,
                first_name=first_name,
                last_name=last_name,
                email=email,
                password=password1,
                is_staff=is_staff,
            )
            registrar_log(request, "Criou usuário", "User", None, username)
            messages.success(request, f'Usuário "{username}" criado com sucesso.')
            return redirect('simulador:usuarios_lista')

    form_data = {'first_name': request.POST.get('first_name', ''), 'last_name': request.POST.get('last_name', ''), 'email': request.POST.get('email', ''), 'is_staff': request.POST.get('is_staff') == 'on'}
    return render(request, 'simulador/usuario_form.html', {'modo': 'criar', 'usuario': None, 'form_data': form_data})


@staff_required
@require_http_methods(["GET", "POST"])
def usuario_editar(request, pk):
    usuario = get_object_or_404(User, pk=pk)

    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        is_staff = request.POST.get('is_staff') == 'on'
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')

        errors = []
        if password1:
            if password1 != password2:
                errors.append('As senhas não conferem.')
            elif len(password1) < 6:
                errors.append('A senha deve ter no mínimo 6 caracteres.')

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            usuario.first_name = first_name
            usuario.last_name = last_name
            usuario.email = email
            if usuario.pk != request.user.pk:
                usuario.is_staff = is_staff
            if password1:
                usuario.set_password(password1)
            usuario.save()
            registrar_log(request, "Editou usuário", "User", usuario.pk, usuario.username)
            messages.success(request, f'Usuário "{usuario.username}" atualizado com sucesso.')
            return redirect('simulador:usuarios_lista')

    form_data = {'first_name': request.POST.get('first_name', usuario.first_name), 'last_name': request.POST.get('last_name', usuario.last_name), 'email': request.POST.get('email', usuario.email), 'is_staff': (request.POST.get('is_staff') == 'on') if request.method == 'POST' else usuario.is_staff}
    return render(request, 'simulador/usuario_form.html', {'modo': 'editar', 'usuario': usuario, 'form_data': form_data})


@staff_required
@require_POST
def usuario_toggle_ativo(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    if usuario.pk == request.user.pk:
        messages.error(request, 'Você não pode desativar sua própria conta.')
    else:
        usuario.is_active = not usuario.is_active
        usuario.save(update_fields=['is_active'])
        estado = 'ativado' if usuario.is_active else 'desativado'
        registrar_log(request, f"Usuário {estado}", "User", usuario.pk, usuario.username)
        messages.success(request, f'Usuário "{usuario.username}" {estado} com sucesso.')
    return redirect('simulador:usuarios_lista')


# ── BCB API helper ─────────────────────────────────────────────────────────────

_bcb_cache = {}

def _bcb_fetch(codigo):
    """Busca o último valor de uma série do SGS/BCB com cache de 1h."""
    agora = datetime.datetime.now()
    if codigo in _bcb_cache:
        valor, ts = _bcb_cache[codigo]
        if (agora - ts).seconds < 3600:
            return valor
    try:
        url = f'https://api.bcb.gov.br/dados/serie/bcdata.sgs.{codigo}/dados/ultimos/1?formato=json'
        with urllib.request.urlopen(url, timeout=4) as resp:
            dados = json.loads(resp.read().decode())
            valor = float(dados[0]['valor'].replace(',', '.'))
            _bcb_cache[codigo] = (valor, agora)
            return valor
    except Exception:
        return None


# ── CET — Custo Efetivo Total ──────────────────────────────────────────────────

@login_required
@require_http_methods(["GET", "POST"])
def cet(request):
    resultado = None
    if request.method == 'POST':
        try:
            valor_financiado = float(request.POST.get('valor_financiado', 0))
            taxa_juros = float(request.POST.get('taxa_juros', 0))
            prazo_meses = int(request.POST.get('meses', 0))
            sistema = request.POST.get('sistema', 'SAC').upper()
            mip_mensal = float(request.POST.get('mip_mensal', 0) or 0)
            dfi_mensal = float(request.POST.get('dfi_mensal', 0) or 0)
            tarifa_emissao = float(request.POST.get('tarifa_emissao', 0) or 0)
            tarifa_avaliacao = float(request.POST.get('tarifa_avaliacao', 0) or 0)
            valor_imovel = float(request.POST.get('valor_imovel', 0) or valor_financiado)

            if valor_financiado <= 0 or taxa_juros <= 0 or prazo_meses <= 0:
                raise ValueError("Preencha os campos obrigatórios com valores positivos.")

            fn = calcular_sac if sistema == 'SAC' else calcular_price
            parcelas = fn(valor_financiado, taxa_juros, prazo_meses, mip_mensal, dfi_mensal, valor_imovel)
            fluxos = [float(p['valor']) for p in parcelas]
            tarifas = tarifa_emissao + tarifa_avaliacao
            pv_liq = valor_financiado - tarifas

            # Newton-Raphson: encontrar a taxa r tal que PV = sum(PMT_k/(1+r)^k)
            r = taxa_juros / 100
            for _ in range(200):
                fr = pv_liq - sum(pmt / (1 + r) ** k for k, pmt in enumerate(fluxos, 1))
                dfr = sum(k * pmt / (1 + r) ** (k + 1) for k, pmt in enumerate(fluxos, 1))
                if abs(dfr) < 1e-14:
                    break
                r_new = r - fr / dfr
                if abs(r_new - r) < 1e-10:
                    r = r_new
                    break
                r = max(r_new, 1e-8)

            cet_mensal = r * 100
            cet_anual = ((1 + r) ** 12 - 1) * 100
            taxa_nom_anual = taxa_juros * 12
            cet_anual_nom = cet_mensal * 12
            total_pago = sum(fluxos)
            total_juros = total_pago - valor_financiado
            custo_total_real = total_pago - valor_financiado + tarifas

            resultado = {
                'sistema': sistema,
                'valor_financiado': f'{valor_financiado:.2f}',
                'taxa_juros': f'{taxa_juros:.4f}',
                'taxa_nom_anual': f'{taxa_nom_anual:.2f}',
                'meses': prazo_meses,
                'tarifas': f'{tarifas:.2f}',
                'tarifa_emissao': f'{tarifa_emissao:.2f}',
                'tarifa_avaliacao': f'{tarifa_avaliacao:.2f}',
                'cet_mensal': f'{cet_mensal:.4f}',
                'cet_anual': f'{cet_anual:.2f}',
                'cet_anual_nom': f'{cet_anual_nom:.2f}',
                'diferenca_mensal': f'{cet_mensal - taxa_juros:.4f}',
                'total_pago': f'{total_pago:.2f}',
                'total_juros': f'{total_juros:.2f}',
                'custo_total_real': f'{custo_total_real:.2f}',
                'primera_parcela': parcelas[0]['valor'] if parcelas else '0.00',
            }
        except (ValueError, TypeError) as e:
            messages.error(request, f'Erro: {e}')

    return render(request, 'simulador/cet.html', {'resultado': resultado})


# ── Consórcio vs Financiamento ────────────────────────────────────────────────

@login_required
@require_http_methods(["GET", "POST"])
def consorcio(request):
    resultado = None
    if request.method == 'POST':
        try:
            valor_bem = float(request.POST.get('valor_bem', 0))
            prazo_meses = int(request.POST.get('meses', 0))
            taxa_admin_pct = float(request.POST.get('taxa_admin_pct', 18.0))
            fundo_reserva_pct = float(request.POST.get('fundo_reserva_pct', 3.0))
            taxa_juros_financ = float(request.POST.get('taxa_juros_financ', 0))

            if valor_bem <= 0 or prazo_meses <= 0 or taxa_juros_financ <= 0:
                raise ValueError("Preencha todos os campos com valores positivos.")

            # ── Consórcio ──────────────────────────────────────────────────────
            taxa_admin_total = valor_bem * taxa_admin_pct / 100
            fundo_reserva = valor_bem * fundo_reserva_pct / 100
            total_consorcio = valor_bem + taxa_admin_total + fundo_reserva
            parcela_consorcio = total_consorcio / prazo_meses
            custo_extra_consorcio = taxa_admin_total + fundo_reserva

            # ── Financiamento PRICE ────────────────────────────────────────────
            parcelas_price = calcular_price(valor_bem, taxa_juros_financ, prazo_meses)
            total_financ = sum(float(p['valor']) for p in parcelas_price)
            juros_financ = total_financ - valor_bem
            parcela_financ_1 = float(parcelas_price[0]['valor'])

            diferenca_total = total_consorcio - total_financ
            parcela_menor_consorcio = parcela_consorcio < parcela_financ_1

            passo = max(1, prazo_meses // 60)
            chart_labels = list(range(1, prazo_meses + 1, passo))
            chart_consorcio = [round(parcela_consorcio, 2)] * len(chart_labels)
            chart_financiamento = [float(parcelas_price[i]['valor']) for i in
                                   range(0, prazo_meses, passo) if i < len(parcelas_price)]

            resultado = {
                'valor_bem': f'{valor_bem:.2f}',
                'prazo_meses': prazo_meses,
                'taxa_admin_pct': f'{taxa_admin_pct:.2f}',
                'fundo_reserva_pct': f'{fundo_reserva_pct:.2f}',
                'taxa_juros_financ': f'{taxa_juros_financ:.4f}',
                'taxa_admin_total': f'{taxa_admin_total:.2f}',
                'fundo_reserva': f'{fundo_reserva:.2f}',
                'total_consorcio': f'{total_consorcio:.2f}',
                'parcela_consorcio': f'{parcela_consorcio:.2f}',
                'custo_extra_consorcio': f'{custo_extra_consorcio:.2f}',
                'total_financ': f'{total_financ:.2f}',
                'juros_financ': f'{juros_financ:.2f}',
                'parcela_financ_1': f'{parcela_financ_1:.2f}',
                'diferenca_total': f'{abs(diferenca_total):.2f}',
                'consorcio_mais_barato': total_consorcio < total_financ,
                'parcela_menor_consorcio': parcela_menor_consorcio,
                'chart_labels': json.dumps(chart_labels),
                'chart_consorcio': json.dumps(chart_consorcio),
                'chart_financiamento': json.dumps(chart_financiamento),
            }
        except (ValueError, TypeError) as e:
            messages.error(request, f'Erro: {e}')

    return render(request, 'simulador/consorcio.html', {'resultado': resultado})


# ── Refinanciamento ────────────────────────────────────────────────────────────

@login_required
@require_http_methods(["GET", "POST"])
def refinanciamento(request):
    resultado = None
    if request.method == 'POST':
        try:
            saldo_devedor = float(request.POST.get('saldo_devedor', 0))
            taxa_atual = float(request.POST.get('taxa_atual', 0))
            prazo_restante = int(request.POST.get('prazo_restante', 0))
            taxa_nova = float(request.POST.get('taxa_nova', 0))
            prazo_novo = int(request.POST.get('prazo_novo', 0) or prazo_restante)
            sistema = request.POST.get('sistema', 'PRICE').upper()

            if any(v <= 0 for v in [saldo_devedor, taxa_atual, prazo_restante, taxa_nova]):
                raise ValueError("Preencha todos os campos com valores positivos.")
            if prazo_novo <= 0:
                prazo_novo = prazo_restante

            fn = calcular_sac if sistema == 'SAC' else calcular_price

            parcelas_atual = fn(saldo_devedor, taxa_atual, prazo_restante)
            parcelas_nova = fn(saldo_devedor, taxa_nova, prazo_novo)

            total_atual = sum(float(p['valor']) for p in parcelas_atual)
            total_nova = sum(float(p['valor']) for p in parcelas_nova)
            juros_atual = sum(float(p['juros']) for p in parcelas_atual)
            juros_nova = sum(float(p['juros']) for p in parcelas_nova)

            parcela_atual = float(parcelas_atual[0]['valor'])
            parcela_nova = float(parcelas_nova[0]['valor'])
            economia_mensal = parcela_atual - parcela_nova
            economia_total = total_atual - total_nova

            passo = max(1, max(prazo_restante, prazo_novo) // 60)
            indices_atual = list(range(0, prazo_restante, passo))
            indices_nova = list(range(0, prazo_novo, passo))
            chart_labels = [parcelas_atual[i]['numero'] for i in indices_atual if i < len(parcelas_atual)]
            chart_atual = [float(parcelas_atual[i]['saldo_devedor']) for i in indices_atual if i < len(parcelas_atual)]
            chart_nova_raw = [float(parcelas_nova[i]['saldo_devedor']) for i in indices_nova if i < len(parcelas_nova)]

            if len(chart_nova_raw) < len(chart_labels):
                chart_nova_raw += [0] * (len(chart_labels) - len(chart_nova_raw))

            resultado = {
                'sistema': sistema,
                'saldo_devedor': f'{saldo_devedor:.2f}',
                'taxa_atual': f'{taxa_atual:.4f}',
                'taxa_nova': f'{taxa_nova:.4f}',
                'prazo_restante': prazo_restante,
                'prazo_novo': prazo_novo,
                'parcela_atual': f'{parcela_atual:.2f}',
                'parcela_nova': f'{parcela_nova:.2f}',
                'economia_mensal': f'{economia_mensal:.2f}',
                'economia_total': f'{economia_total:.2f}',
                'total_atual': f'{total_atual:.2f}',
                'total_nova': f'{total_nova:.2f}',
                'juros_atual': f'{juros_atual:.2f}',
                'juros_nova': f'{juros_nova:.2f}',
                'economia_juros': f'{juros_atual - juros_nova:.2f}',
                'economia_pct': f'{(economia_total / total_atual * 100):.1f}' if total_atual > 0 else '0',
                'vale_refin': economia_mensal > 0 or economia_total > 0,
                'chart_labels': json.dumps(chart_labels),
                'chart_atual': json.dumps(chart_atual),
                'chart_nova': json.dumps(chart_nova_raw),
            }
        except (ValueError, TypeError) as e:
            messages.error(request, f'Erro: {e}')

    return render(request, 'simulador/refinanciamento.html', {'resultado': resultado})


# ── Relatório Gerencial PDF ────────────────────────────────────────────────────

@staff_required
def relatorio_pdf(request):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    qs = Simulation.objects.select_related('usuario').all()
    total = qs.count()
    volume = float(qs.aggregate(total=Sum('valor_imovel'))['total'] or 0)
    aprovados = qs.filter(status='aprovado').count()
    ticket_medio = volume / total if total > 0 else 0
    sac_count = qs.filter(sistema='SAC').count()
    price_count = qs.filter(sistema='PRICE').count()
    status_qs = qs.values('status').annotate(n=Count('status'))
    status_map = dict(Simulation.STATUS_CHOICES)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    azul = colors.HexColor('#1a3c5e')
    azul_claro = colors.HexColor('#2e86c1')
    cinza = colors.HexColor('#f0f4f8')

    titulo_style = ParagraphStyle('titulo', parent=styles['Title'],
                                   textColor=azul, fontSize=20, spaceAfter=4)
    sub_style = ParagraphStyle('sub', parent=styles['Normal'],
                                textColor=azul_claro, fontSize=10, spaceAfter=12)
    h2_style = ParagraphStyle('h2', parent=styles['Heading2'],
                               textColor=azul, fontSize=13, spaceBefore=16, spaceAfter=6)

    story = [
        Paragraph('Relatório Gerencial', titulo_style),
        Paragraph(f'Emitido em {datetime.date.today().strftime("%d/%m/%Y")} por {request.user.get_full_name() or request.user.username}', sub_style),
        HRFlowable(width='100%', thickness=1, color=azul_claro, spaceAfter=12),
        Spacer(1, 0.2*cm),
    ]

    story.append(Paragraph('Indicadores Gerais', h2_style))
    kpis = [
        ['Métrica', 'Valor'],
        ['Total de Simulações', str(total)],
        ['Volume Total (Imóveis)', f'R$ {volume:,.2f}'],
        ['Ticket Médio', f'R$ {ticket_medio:,.2f}'],
        ['Aprovadas', f'{aprovados} ({(aprovados/total*100):.1f}%)' if total else '0'],
        ['Sistema SAC', str(sac_count)],
        ['Sistema PRICE', str(price_count)],
    ]
    t_kpi = Table(kpis, colWidths=[9*cm, 9*cm])
    t_kpi.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), azul),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 1), (0, -1), cinza),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 1), (0, -1), azul),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('PADDING', (0, 0), (-1, -1), 7),
    ]))
    story += [t_kpi, Spacer(1, 0.4*cm)]

    story.append(Paragraph('Distribuição por Status', h2_style))
    status_rows = [['Status', 'Quantidade', '% do Total']]
    for s in status_qs.order_by('-n'):
        pct = s['n'] / total * 100 if total else 0
        status_rows.append([status_map.get(s['status'], s['status']), str(s['n']), f'{pct:.1f}%'])
    t_status = Table(status_rows, colWidths=[8*cm, 5*cm, 5*cm])
    t_status.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), azul),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f7ff')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story += [t_status, Spacer(1, 0.4*cm)]

    story.append(Paragraph('Últimas 20 Simulações', h2_style))
    recentes = qs.order_by('-criado_em')[:20]
    cab = [['Cliente', 'Usuário', 'Imóvel (R$)', 'Sistema', 'Status', 'Data']]
    linhas = cab + [
        [s.cliente, s.usuario.username,
         f'{float(s.valor_imovel):,.0f}', s.sistema,
         status_map.get(s.status, s.status),
         s.criado_em.strftime('%d/%m/%Y')]
        for s in recentes
    ]
    t_rec = Table(linhas, colWidths=[4*cm, 3*cm, 3.5*cm, 2*cm, 2.5*cm, 2.5*cm])
    t_rec.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), azul),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f7ff')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#dee2e6')),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(t_rec)

    doc.build(story)
    buffer.seek(0)
    nome = f'relatorio_gerencial_{datetime.date.today().strftime("%Y%m%d")}.pdf'
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nome}"'
    return response


# ── Taxas BCB ──────────────────────────────────────────────────────────────────

@login_required
def taxas_bcb(request):
    selic = _bcb_fetch(432)
    ipca = _bcb_fetch(13522)
    cdi = _bcb_fetch(4389)
    from django.http import JsonResponse
    return JsonResponse({
        'selic': selic,
        'ipca': ipca,
        'cdi': cdi,
        'atualizado_em': datetime.datetime.now().strftime('%H:%M:%S'),
    })


# ── API REST ────────────────────────────────────────────────────────────────────

@login_required
@require_POST
def api_simular(request):
    from django.http import JsonResponse
    try:
        data = json.loads(request.body)
        valor_financiado = float(data.get('valor_financiado', 0))
        taxa_juros = float(data.get('taxa_juros', 0))
        prazo_meses = int(data.get('prazo_meses', 0))
        sistema = data.get('sistema', 'SAC').upper()
        mip_mensal = float(data.get('mip_mensal', 0))
        dfi_mensal = float(data.get('dfi_mensal', 0))
        valor_imovel = float(data.get('valor_imovel', valor_financiado))

        if valor_financiado <= 0 or taxa_juros <= 0 or prazo_meses <= 0:
            return JsonResponse({'erro': 'Parâmetros inválidos.'}, status=400)

        fn = calcular_sac if sistema == 'SAC' else calcular_price
        parcelas = fn(valor_financiado, taxa_juros, prazo_meses, mip_mensal, dfi_mensal, valor_imovel)
        total_pago = sum(float(p['valor']) for p in parcelas)

        return JsonResponse({
            'sistema': sistema,
            'valor_financiado': valor_financiado,
            'taxa_juros_mensal': taxa_juros,
            'prazo_meses': prazo_meses,
            'primeira_parcela': parcelas[0]['valor'] if parcelas else '0.00',
            'ultima_parcela': parcelas[-1]['valor'] if parcelas else '0.00',
            'total_pago': f'{total_pago:.2f}',
            'total_juros': f'{total_pago - valor_financiado:.2f}',
            'parcelas': parcelas,
        })
    except (ValueError, TypeError, json.JSONDecodeError) as e:
        return JsonResponse({'erro': str(e)}, status=400)


@login_required
@require_POST
def api_oraculo(request):
    from django.http import JsonResponse
    try:
        data = json.loads(request.body)
        renda = float(data.get('renda', 0))
        entrada = float(data.get('entrada', 0))
        prazo_anos = int(data.get('prazo_anos', 30))
        taxa_anual = float(data.get('taxa_anual', 9.99))
        comprometimento = float(data.get('comprometimento', 30))

        if renda <= 0:
            return JsonResponse({'erro': 'Renda deve ser maior que zero.'}, status=400)

        margem_parcela = renda * (comprometimento / 100)
        taxa_mensal = (taxa_anual / 100) / 12
        meses = prazo_anos * 12
        valor_financiavel = margem_parcela * ((1 - (1 + taxa_mensal) ** (-meses)) / taxa_mensal)
        poder_compra = valor_financiavel + entrada

        return JsonResponse({
            'poder_compra': round(poder_compra, 2),
            'valor_financiavel': round(valor_financiavel, 2),
            'margem_parcela': round(margem_parcela, 2),
        })
    except (ValueError, TypeError, json.JSONDecodeError) as e:
        return JsonResponse({'erro': str(e)}, status=400)


# ── Auditoria helper ───────────────────────────────────────────────────────────

def registrar_log(request, acao, obj_tipo='', obj_id=None, descricao=''):
    try:
        AuditLog.objects.create(
            usuario=request.user if request.user.is_authenticated else None,
            acao=acao,
            objeto_tipo=obj_tipo,
            objeto_id=obj_id,
            descricao=descricao,
        )
    except Exception:
        pass


# ── Grupo A: Novas ferramentas ─────────────────────────────────────────────────

BANCOS_REFERENCIA = [
    ('Caixa Econômica', 8.99),
    ('Banco do Brasil', 9.49),
    ('Bradesco', 9.99),
    ('Itaú', 10.49),
    ('Santander', 10.99),
]


@login_required
@require_http_methods(["GET", "POST"])
def comparativo_bancos(request):
    resultado = None
    if request.method == "POST":
        try:
            valor_financiado = float(request.POST.get("valor_financiado", 0))
            prazo_meses = int(request.POST.get("prazo_meses", 0))
            sistema = request.POST.get("sistema", "SAC").upper()
            if valor_financiado <= 0 or prazo_meses <= 0:
                raise ValueError("Valores inválidos.")
            fn = calcular_sac if sistema == "SAC" else calcular_price
            bancos = []
            for nome, taxa_anual in BANCOS_REFERENCIA:
                taxa_mensal = round(taxa_anual / 12, 4)
                parcelas = fn(valor_financiado, taxa_mensal, prazo_meses, 0, 0, valor_financiado)
                total_pago = sum(float(p["valor"]) for p in parcelas)
                bancos.append({
                    "nome": nome,
                    "taxa_anual": taxa_anual,
                    "taxa_mensal": round(taxa_mensal, 4),
                    "primeira_parcela": float(parcelas[0]["valor"]),
                    "total_pago": total_pago,
                    "total_juros": total_pago - valor_financiado,
                })
            melhor = min(bancos, key=lambda b: b["total_pago"])
            resultado = {
                "bancos": bancos,
                "melhor": melhor,
                "valor_financiado": valor_financiado,
                "prazo_meses": prazo_meses,
                "sistema": sistema,
                "chart_labels": json.dumps([b["nome"] for b in bancos]),
                "chart_primeiras": json.dumps([round(b["primeira_parcela"], 2) for b in bancos]),
                "chart_totais": json.dumps([round(b["total_pago"], 2) for b in bancos]),
            }
        except (ValueError, TypeError) as e:
            messages.error(request, f"Erro: {e}")
    return render(request, "simulador/comparativo_bancos.html", {"resultado": resultado})


@login_required
@require_http_methods(["GET", "POST"])
def mcmv(request):
    resultado = None
    if request.method == "POST":
        try:
            renda = float(request.POST.get("renda", 0))
            valor_imovel = float(request.POST.get("valor_imovel", 0))
            entrada = float(request.POST.get("entrada", 0))
            prazo_anos = int(request.POST.get("prazo_anos", 30))

            if renda <= 0 or valor_imovel <= 0:
                raise ValueError("Renda e valor do imóvel devem ser maiores que zero.")
            if entrada >= valor_imovel:
                raise ValueError("Entrada não pode ser maior ou igual ao valor do imóvel.")

            if renda <= 2640:
                faixa = "Faixa 1"
                subsidio = 55000.0
                taxa_anual = 4.25
            elif renda <= 4400:
                faixa = "Faixa 1,5"
                subsidio = 47500.0
                taxa_anual = 4.75
            elif renda <= 8000:
                faixa = "Faixa 2"
                subsidio = 29000.0 * (8000.0 - renda) / (8000.0 - 4400.0)
                subsidio = max(0.0, subsidio)
                taxa_anual = 7.66
            elif renda <= 12000:
                faixa = "Faixa 3"
                subsidio = 0.0
                taxa_anual = 9.99
            else:
                faixa = "Fora do MCMV"
                subsidio = 0.0
                taxa_anual = 9.99

            subsidio_efetivo = min(subsidio, max(0, valor_imovel - entrada))
            valor_financiado = valor_imovel - entrada - subsidio_efetivo
            if valor_financiado <= 0:
                raise ValueError("Com o subsídio o imóvel pode ser quitado sem financiamento.")

            prazo_meses = prazo_anos * 12
            taxa_mensal = round(taxa_anual / 12, 4)
            parcelas = calcular_price(valor_financiado, taxa_mensal, prazo_meses, 0, 0, valor_financiado)
            total_pago = sum(float(p["valor"]) for p in parcelas)

            resultado = {
                "faixa": faixa,
                "subsidio": subsidio_efetivo,
                "taxa_anual": taxa_anual,
                "valor_financiado": valor_financiado,
                "primeira_parcela": float(parcelas[0]["valor"]),
                "total_pago": total_pago,
                "renda": renda,
                "valor_imovel": valor_imovel,
                "entrada": entrada,
                "prazo_anos": prazo_anos,
            }
        except (ValueError, TypeError) as e:
            messages.error(request, f"Erro: {e}")
    return render(request, "simulador/mcmv.html", {"resultado": resultado})


@login_required
@require_http_methods(["GET", "POST"])
def renda_minima(request):
    resultado = None
    if request.method == "POST":
        try:
            valor_imovel = float(request.POST.get("valor_imovel", 0))
            entrada = float(request.POST.get("entrada", 0))
            taxa_juros = float(request.POST.get("taxa_juros", 0))
            prazo_meses = int(request.POST.get("prazo_meses", 0))
            comprometimento = float(request.POST.get("comprometimento", 30))

            if valor_imovel <= 0 or taxa_juros <= 0 or prazo_meses <= 0:
                raise ValueError("Valores inválidos.")
            if entrada >= valor_imovel:
                raise ValueError("Entrada não pode ser maior ou igual ao valor do imóvel.")
            if not (10 <= comprometimento <= 50):
                raise ValueError("Comprometimento deve estar entre 10% e 50%.")

            valor_financiado = valor_imovel - entrada
            parcelas_price = calcular_price(valor_financiado, taxa_juros, prazo_meses, 0, 0, valor_financiado)
            primeira_parcela = float(parcelas_price[0]["valor"])

            cenarios = []
            for comp in [25, 30, 35]:
                renda_min = primeira_parcela / (comp / 100)
                cenarios.append({"comprometimento": comp, "renda_minima": renda_min})

            renda_calc = primeira_parcela / (comprometimento / 100)

            resultado = {
                "renda_minima": renda_calc,
                "primeira_parcela": primeira_parcela,
                "comprometimento": comprometimento,
                "cenarios": cenarios,
                "valor_imovel": valor_imovel,
                "entrada": entrada,
                "valor_financiado": valor_financiado,
                "taxa_juros": taxa_juros,
                "prazo_meses": prazo_meses,
            }
        except (ValueError, TypeError) as e:
            messages.error(request, f"Erro: {e}")
    return render(request, "simulador/renda_minima.html", {"resultado": resultado})


@login_required
@require_http_methods(["GET", "POST"])
def prazo_idade(request):
    resultado = None
    if request.method == "POST":
        try:
            idade = int(request.POST.get("idade", 0))
            prazo_desejado = int(request.POST.get("prazo_desejado", 0))

            if not (18 <= idade <= 80):
                raise ValueError("Idade deve estar entre 18 e 80 anos.")
            if prazo_desejado <= 0:
                raise ValueError("Prazo desejado deve ser maior que zero.")

            meses_restantes_80 = (80 * 12 + 6) - (idade * 12)
            prazo_max = max(0, meses_restantes_80)
            prazo_efetivo = min(prazo_desejado, prazo_max)
            foi_limitado = prazo_efetivo < prazo_desejado
            idade_final = idade + (prazo_efetivo // 12)
            meses_finais = prazo_efetivo % 12

            resultado = {
                "idade": idade,
                "prazo_desejado": prazo_desejado,
                "prazo_max": prazo_max,
                "prazo_efetivo": prazo_efetivo,
                "foi_limitado": foi_limitado,
                "diferenca": prazo_desejado - prazo_efetivo,
                "idade_ao_final": idade_final,
                "meses_ao_final": meses_finais,
            }
        except (ValueError, TypeError) as e:
            messages.error(request, f"Erro: {e}")
    return render(request, "simulador/prazo_idade.html", {"resultado": resultado})


@login_required
@require_http_methods(["GET", "POST"])
def financiamento_ipca(request):
    resultado = None
    if request.method == "POST":
        try:
            valor_financiado = float(request.POST.get("valor_financiado", 0))
            prazo_meses = int(request.POST.get("prazo_meses", 0))
            spread = float(request.POST.get("spread", 3.5))
            ipca_projetado = float(request.POST.get("ipca_projetado", 4.5))

            if valor_financiado <= 0 or prazo_meses <= 0:
                raise ValueError("Valores inválidos.")

            def calc_cenario(ipca_anual):
                taxa_efetiva_anual = ((1 + ipca_anual / 100) * (1 + spread / 100)) - 1
                taxa_mensal = ((1 + taxa_efetiva_anual) ** (1 / 12) - 1) * 100
                parcelas = calcular_sac(valor_financiado, round(taxa_mensal, 4), prazo_meses, 0, 0, valor_financiado)
                total_pago = sum(float(p["valor"]) for p in parcelas)
                return {
                    "taxa_anual": round(taxa_efetiva_anual * 100, 2),
                    "taxa_mensal": round(taxa_mensal, 4),
                    "primeira_parcela": float(parcelas[0]["valor"]),
                    "ultima_parcela": float(parcelas[-1]["valor"]),
                    "total_pago": total_pago,
                    "total_juros": total_pago - valor_financiado,
                }

            cenario_base = calc_cenario(ipca_projetado)
            cenario_pess = calc_cenario(ipca_projetado + 2)
            cenario_otim = calc_cenario(max(0, ipca_projetado - 2))

            resultado = {
                "valor_financiado": valor_financiado,
                "prazo_meses": prazo_meses,
                "spread": spread,
                "ipca_projetado": ipca_projetado,
                "base": cenario_base,
                "pessimista": cenario_pess,
                "otimista": cenario_otim,
                "chart_labels": json.dumps(["Otimista", "Base", "Pessimista"]),
                "chart_primeiras": json.dumps([
                    round(cenario_otim["primeira_parcela"], 2),
                    round(cenario_base["primeira_parcela"], 2),
                    round(cenario_pess["primeira_parcela"], 2),
                ]),
                "chart_totais": json.dumps([
                    round(cenario_otim["total_pago"], 2),
                    round(cenario_base["total_pago"], 2),
                    round(cenario_pess["total_pago"], 2),
                ]),
            }
        except (ValueError, TypeError) as e:
            messages.error(request, f"Erro: {e}")
    return render(request, "simulador/financiamento_ipca.html", {"resultado": resultado})


# ── Grupo B: Clientes ──────────────────────────────────────────────────────────

@login_required
def clientes_lista(request):
    q = request.GET.get("q", "").strip()
    qs = Cliente.objects.filter(usuario=request.user)
    if q:
        qs = qs.filter(nome__icontains=q)
    paginator = Paginator(qs, 20)
    page = paginator.get_page(request.GET.get("page"))
    return render(request, "simulador/clientes_lista.html", {"page_obj": page, "q": q})


@login_required
@require_http_methods(["GET", "POST"])
def cliente_criar(request):
    if request.method == "POST":
        nome = request.POST.get("nome", "").strip()
        email = request.POST.get("email", "").strip()
        telefone = request.POST.get("telefone", "").strip()
        cpf = request.POST.get("cpf", "").strip()
        renda_raw = request.POST.get("renda_mensal", "").strip()
        observacoes = request.POST.get("observacoes", "").strip()
        if not nome:
            messages.error(request, "O nome do cliente é obrigatório.")
            return render(request, "simulador/cliente_form.html", {"acao": "Novo", "form_data": request.POST})
        try:
            renda_mensal = float(renda_raw.replace(",", ".")) if renda_raw else None
        except ValueError:
            renda_mensal = None
        c = Cliente.objects.create(
            usuario=request.user,
            nome=nome, email=email, telefone=telefone, cpf=cpf,
            renda_mensal=renda_mensal, observacoes=observacoes,
        )
        registrar_log(request, "Criou cliente", "Cliente", c.pk, nome)
        messages.success(request, f"Cliente {nome} cadastrado com sucesso.")
        return redirect("simulador:cliente_detalhe", pk=c.pk)
    return render(request, "simulador/cliente_form.html", {"acao": "Novo"})


@login_required
@require_http_methods(["GET", "POST"])
def cliente_editar(request, pk):
    c = get_object_or_404(Cliente, pk=pk, usuario=request.user)
    if request.method == "POST":
        nome = request.POST.get("nome", "").strip()
        if not nome:
            messages.error(request, "O nome do cliente é obrigatório.")
            return render(request, "simulador/cliente_form.html", {"acao": "Editar", "cliente": c})
        renda_raw = request.POST.get("renda_mensal", "").strip()
        try:
            renda_mensal = float(renda_raw.replace(",", ".")) if renda_raw else None
        except ValueError:
            renda_mensal = None
        c.nome = nome
        c.email = request.POST.get("email", "").strip()
        c.telefone = request.POST.get("telefone", "").strip()
        c.cpf = request.POST.get("cpf", "").strip()
        c.renda_mensal = renda_mensal
        c.observacoes = request.POST.get("observacoes", "").strip()
        c.save()
        registrar_log(request, "Editou cliente", "Cliente", c.pk, nome)
        messages.success(request, "Cliente atualizado.")
        return redirect("simulador:cliente_detalhe", pk=c.pk)
    return render(request, "simulador/cliente_form.html", {"acao": "Editar", "cliente": c})


@login_required
@require_POST
def cliente_excluir(request, pk):
    c = get_object_or_404(Cliente, pk=pk, usuario=request.user)
    nome = c.nome
    c.delete()
    registrar_log(request, "Excluiu cliente", "Cliente", pk, nome)
    messages.success(request, f"Cliente {nome} excluído.")
    return redirect("simulador:clientes_lista")


@login_required
def cliente_detalhe(request, pk):
    c = get_object_or_404(Cliente, pk=pk, usuario=request.user)
    simulacoes = c.simulacoes_obj.order_by("-criado_em")
    return render(request, "simulador/cliente_detalhe.html", {"cliente": c, "simulacoes": simulacoes})


# ── Pipeline Kanban ────────────────────────────────────────────────────────────

@login_required
def pipeline(request):
    qs = Simulation.objects.filter(usuario=request.user) if not request.user.is_staff else Simulation.objects.all()
    colunas = {
        "novo": {"label": "Novo", "cor": "secondary", "icone": "bi-inbox", "cards": []},
        "em_analise": {"label": "Em Análise", "cor": "warning", "icone": "bi-hourglass-split", "cards": []},
        "aprovado": {"label": "Aprovado", "cor": "success", "icone": "bi-check-circle", "cards": []},
        "reprovado": {"label": "Reprovado", "cor": "danger", "icone": "bi-x-circle", "cards": []},
    }
    for sim in qs.select_related("usuario"):
        if sim.status in colunas:
            colunas[sim.status]["cards"].append(sim)
    return render(request, "simulador/pipeline.html", {"colunas": colunas})


@login_required
@require_POST
def mover_card(request, pk):
    from django.http import JsonResponse
    if request.user.is_staff:
        sim = get_object_or_404(Simulation, pk=pk)
    else:
        sim = get_object_or_404(Simulation, pk=pk, usuario=request.user)
    try:
        data = json.loads(request.body)
        novo_status = data.get("status", "")
    except (json.JSONDecodeError, AttributeError):
        novo_status = request.POST.get("status", "")
    status_validos = [s for s, _ in Simulation.STATUS_CHOICES]
    if novo_status not in status_validos:
        return JsonResponse({"erro": "Status inválido."}, status=400)
    antigo = sim.status
    sim.status = novo_status
    sim.save(update_fields=["status"])
    registrar_log(request, "Moveu card no pipeline", "Simulation", pk,
                  f"{sim.cliente}: {antigo} → {novo_status}")
    return JsonResponse({"ok": True, "status": novo_status})


# ── Metas do corretor ──────────────────────────────────────────────────────────

@login_required
def metas(request):
    hoje = datetime.date.today()
    meta_atual = MetaCorretor.objects.filter(
        usuario=request.user, mes=hoje.month, ano=hoje.year
    ).first()
    total_sims_mes = Simulation.objects.filter(
        usuario=request.user,
        criado_em__month=hoje.month,
        criado_em__year=hoje.year,
    ).count()
    valor_mes = Simulation.objects.filter(
        usuario=request.user,
        criado_em__month=hoje.month,
        criado_em__year=hoje.year,
    ).aggregate(total=Sum("valor_imovel"))["total"] or 0

    progresso_sims = 0
    progresso_valor = 0
    if meta_atual:
        if meta_atual.meta_simulacoes > 0:
            progresso_sims = min(100, int(total_sims_mes / meta_atual.meta_simulacoes * 100))
        if meta_atual.meta_valor > 0:
            progresso_valor = min(100, int(float(valor_mes) / float(meta_atual.meta_valor) * 100))

    historico = MetaCorretor.objects.filter(usuario=request.user).order_by("-ano", "-mes")[:12]
    return render(request, "simulador/metas.html", {
        "meta_atual": meta_atual,
        "total_sims_mes": total_sims_mes,
        "valor_mes": valor_mes,
        "progresso_sims": progresso_sims,
        "progresso_valor": progresso_valor,
        "historico": historico,
        "mes_atual": hoje.month,
        "ano_atual": hoje.year,
        "meses": ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"],
    })


@login_required
@require_http_methods(["GET", "POST"])
def meta_criar(request):
    hoje = datetime.date.today()
    if request.method == "POST":
        try:
            mes = int(request.POST.get("mes", hoje.month))
            ano = int(request.POST.get("ano", hoje.year))
            meta_simulacoes = int(request.POST.get("meta_simulacoes", 0))
            val_raw = request.POST.get("meta_valor", "0")
            meta_valor = float(str(val_raw).replace(".", "").replace(",", ".")) if val_raw else 0
            if MetaCorretor.objects.filter(usuario=request.user, mes=mes, ano=ano).exists():
                messages.error(request, f"Já existe uma meta para {mes:02d}/{ano}.")
                return redirect("simulador:metas")
            MetaCorretor.objects.create(
                usuario=request.user, mes=mes, ano=ano,
                meta_simulacoes=meta_simulacoes, meta_valor=meta_valor,
            )
            messages.success(request, "Meta criada com sucesso.")
            return redirect("simulador:metas")
        except (ValueError, TypeError) as e:
            messages.error(request, f"Erro: {e}")
    anos = list(range(hoje.year - 1, hoje.year + 3))
    return render(request, "simulador/meta_form.html", {
        "acao": "Nova",
        "mes_padrao": hoje.month,
        "ano_padrao": hoje.year,
        "anos": anos,
    })


@login_required
@require_http_methods(["GET", "POST"])
def meta_editar(request, pk):
    meta = get_object_or_404(MetaCorretor, pk=pk, usuario=request.user)
    if request.method == "POST":
        try:
            meta.meta_simulacoes = int(request.POST.get("meta_simulacoes", 0))
            val_raw = request.POST.get("meta_valor", "0")
            meta.meta_valor = float(str(val_raw).replace(".", "").replace(",", ".")) if val_raw else 0
            meta.save()
            messages.success(request, "Meta atualizada.")
            return redirect("simulador:metas")
        except (ValueError, TypeError) as e:
            messages.error(request, f"Erro: {e}")
    return render(request, "simulador/meta_form.html", {"acao": "Editar", "meta": meta})


@login_required
@require_POST
def meta_excluir(request, pk):
    meta = get_object_or_404(MetaCorretor, pk=pk, usuario=request.user)
    meta.delete()
    messages.success(request, "Meta excluída.")
    return redirect("simulador:metas")


# ── Log de auditoria ───────────────────────────────────────────────────────────

@staff_required
def logs_auditoria(request):
    qs = AuditLog.objects.select_related("usuario").all()
    filtro_usuario = request.GET.get("usuario", "").strip()
    filtro_data = request.GET.get("data", "").strip()
    filtro_acao = request.GET.get("acao", "").strip()
    if filtro_usuario:
        qs = qs.filter(usuario__username__icontains=filtro_usuario)
    if filtro_data:
        try:
            dt = datetime.datetime.strptime(filtro_data, "%Y-%m-%d").date()
            qs = qs.filter(criado_em__date=dt)
        except ValueError:
            pass
    if filtro_acao:
        qs = qs.filter(acao__icontains=filtro_acao)
    paginator = Paginator(qs, 50)
    page = paginator.get_page(request.GET.get("page"))
    return render(request, "simulador/logs_auditoria.html", {
        "page_obj": page,
        "filtro_usuario": filtro_usuario,
        "filtro_data": filtro_data,
        "filtro_acao": filtro_acao,
    })


# ── Grupo C: Relatório por corretor ───────────────────────────────────────────

@staff_required
def relatorio_corretores(request):
    usuarios = User.objects.filter(is_active=True).annotate(
        total_sims=Count("simulacoes"),
        total_aprovado=Count("simulacoes", filter=Q(simulacoes__status="aprovado")),
        total_reprovado=Count("simulacoes", filter=Q(simulacoes__status="reprovado")),
        valor_total=Sum("simulacoes__valor_imovel"),
    ).order_by("-total_sims")
    return render(request, "simulador/relatorio_corretores.html", {"usuarios": usuarios})


# ── Grupo D: 2FA ───────────────────────────────────────────────────────────────

@login_required
@require_http_methods(["GET", "POST"])
def setup_2fa(request):
    import pyotp
    profile, _ = UserProfile.objects.get_or_create(user=request.user)

    if request.method == "POST":
        acao = request.POST.get("acao", "")

        if acao == "ativar":
            secret = request.POST.get("secret", "").strip()
            codigo = request.POST.get("codigo", "").strip()
            if not secret or not codigo:
                messages.error(request, "Código inválido.")
                return redirect("simulador:setup_2fa")
            totp = pyotp.TOTP(secret)
            if totp.verify(codigo, valid_window=1):
                profile.totp_secret = secret
                profile.totp_enabled = True
                profile.save()
                request.session["_2fa_done"] = True
                registrar_log(request, "Ativou 2FA", "UserProfile", profile.pk)
                messages.success(request, "Autenticação de dois fatores ativada com sucesso!")
                return redirect("simulador:perfil")
            else:
                messages.error(request, "Código incorreto. Tente novamente.")
                return redirect("simulador:setup_2fa")

        elif acao == "desativar":
            profile.totp_enabled = False
            profile.totp_secret = ""
            profile.save()
            registrar_log(request, "Desativou 2FA", "UserProfile", profile.pk)
            messages.success(request, "2FA desativado.")
            return redirect("simulador:perfil")

    novo_secret = pyotp.random_base32()
    totp = pyotp.TOTP(novo_secret)
    uri = totp.provisioning_uri(
        name=request.user.email or request.user.username,
        issuer_name="Imobiliária"
    )
    return render(request, "simulador/setup_2fa.html", {
        "profile": profile,
        "secret": novo_secret,
        "uri": uri,
    })


def verificar_2fa(request):
    import pyotp
    if not request.user.is_authenticated:
        return redirect("login")
    try:
        profile = request.user.profile
    except UserProfile.DoesNotExist:
        return redirect("simulador:simular")

    if not profile.totp_enabled:
        request.session["_2fa_done"] = True
        return redirect("simulador:simular")

    if request.session.get("_2fa_done"):
        return redirect("simulador:simular")

    if request.method == "POST":
        codigo = request.POST.get("codigo", "").strip()
        totp = pyotp.TOTP(profile.totp_secret)
        if totp.verify(codigo, valid_window=1):
            request.session["_2fa_done"] = True
            next_url = request.GET.get("next") or "/"
            return redirect(next_url)
        else:
            messages.error(request, "Código incorreto.")
    return render(request, "simulador/verificar_2fa.html")
